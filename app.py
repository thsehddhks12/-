import streamlit as st
import pandas as pd
import openpyxl

st.set_page_config(page_title="생산실적 요약 대시보드", layout="wide")

file_path = "▣ 26년 월간 실적보고 (7월).xlsx"

# 1. 엑셀 데이터 추출
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

report_title = ws.cell(row=2, column=2).value or "26년 07월 생산실적 보고"

def safe_val(row, col):
    val = ws.cell(row=row, column=col).value
    if val is None or str(val).startswith('='):
        return 0
    try:
        return float(val)
    except:
        return 0

# KPI 수치 추출
order_qty = safe_val(7, 15) or safe_val(7, 14)
plan_qty = safe_val(8, 15) or safe_val(8, 14)
capa_qty = safe_val(9, 15) or safe_val(9, 14)
actual_qty = safe_val(10, 15) or safe_val(10, 14)
achieve_rate = (actual_qty / plan_qty * 100) if plan_qty > 0 else 0

ship_plan = safe_val(50, 11)
ship_actual = safe_val(50, 13)

# ----------------------------------------------------
# UI 메인 헤더 & KPI 카드
# ----------------------------------------------------
st.title(f"📊 [임원 보고용] {report_title} 요약 대시보드")
st.markdown("---")

st.subheader("1. 핵심 생산 KPI & 출고 요약")
k1, k2, k3, k4, k5 = st.columns(5)
k1.metric("발주량", f"{order_qty:,.0f} EA")
k2.metric("생산계획", f"{plan_qty:,.0f} EA")
k3.metric("CAPA계획", f"{capa_qty:,.0f} EA")
k4.metric("생산실적", f"{actual_qty:,.0f} EA", delta=f"달성률 {achieve_rate:.1f}%")
k5.metric("완제품출고 (실적/계획)", f"{ship_actual:,.0f} / {ship_plan:,.0f} 건")

st.markdown("---")

# ----------------------------------------------------
# 2. 비가동 요인 종합 분석 (헤더 세부 교정)
# ----------------------------------------------------
st.subheader("2. 비가동 요인 종합 분석")

downtime_headers = []
for c in range(9, 19):
    h20 = ws.cell(row=20, column=c).value
    h21 = ws.cell(row=21, column=c).value
    
    col_name = str(h21).strip() if (h21 and str(h21).strip()) else (str(h20).strip() if (h20 and str(h20).strip()) else f"미지정_{c}")
    downtime_headers.append(col_name)

# 데이터 영역 (22행 ~ 29행)
downtime_rows = []
for r in range(22, 30):
    row_vals = [ws.cell(row=r, column=c).value for c in range(9, 19)]
    if any(v is not None and str(v).strip() != '' for v in row_vals):
        downtime_rows.append(row_vals)

df_downtime = pd.DataFrame(downtime_rows, columns=downtime_headers)

# 불필요한 미지정 공란 열 정제
df_downtime = df_downtime.loc[:, ~df_downtime.columns.str.startswith("미지정_")].fillna("-")

st.dataframe(df_downtime, use_container_width=True)

st.markdown("---")

# ----------------------------------------------------
# 3. 라인별 UPH / PPH 상세 관리 현황 (병합 상위 헤더 전파)
# ----------------------------------------------------
st.subheader("3. 라인별 UPH / PPH 상세 관리 현황")

uph_r29 = [ws.cell(row=29, column=c).value for c in range(2, 18)]
uph_r30 = [ws.cell(row=30, column=c).value for c in range(2, 18)]

uph_cols = []
current_parent = "구분"

for h1, h2 in zip(uph_r29, uph_r30):
    if h1 and str(h1).strip():
        current_parent = str(h1).strip()
    
    sub_name = str(h2).strip() if (h2 and str(h2).strip()) else ""
    full_name = f"{current_parent} ({sub_name})" if sub_name else current_parent
    uph_cols.append(full_name)

# 열 이름 중복 처리
seen = {}
final_uph_cols = []
for col in uph_cols:
    if col in seen:
        seen[col] += 1
        final_uph_cols.append(f"{col} ({seen[col]})")
    else:
        seen[col] = 0
        final_uph_cols.append(col)

uph_rows = []
for r in range(31, 39):
    row_vals = [ws.cell(row=r, column=c).value for c in range(2, 18)]
    if any(v is not None and str(v).strip() != '' for v in row_vals):
        formatted = [round(v, 2) if isinstance(v, float) else (v if v is not None else "-") for v in row_vals]
        uph_rows.append(formatted)

df_uph = pd.DataFrame(uph_rows, columns=final_uph_cols)

st.dataframe(df_uph, use_container_width=True)
