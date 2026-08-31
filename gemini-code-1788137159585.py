import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 1. 원본 엑셀 파일 로드
file_path = "▣ 26년 월간 실적보고 (7월).xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

# ----------------------------------------------------
# 2. 데이터 추출 (Data Extraction)
# ----------------------------------------------------
report_title = ws.cell(row=2, column=2).value or "26년 07월 생산실적 보고"

# (1) 1번 항목: 생산계획 및 실적 (Column O = 15번째 열)
order_qty = ws.cell(row=7, column=15).value or 0    # 7행 O열: 발주량
plan_qty = ws.cell(row=8, column=15).value or 0     # 8행 O열: 생산계획수량
capa_qty = ws.cell(row=9, column=15).value or 0     # 9행 O열: capa 계획 수량
actual_qty = ws.cell(row=10, column=15).value or 0  # 10행 O열: 생산실적 수량
achieve_rate = (actual_qty / plan_qty * 100) if plan_qty > 0 else 0

# (2) 투입시간 (19행 F열=6, K열=11) 및 완제품 출고 (50행 K열=11, M열=13)
plan_time = ws.cell(row=19, column=6).value or 0
actual_time = ws.cell(row=19, column=11).value or 0
ship_plan = ws.cell(row=50, column=11).value or 0
ship_actual = ws.cell(row=50, column=13).value or 0

# (3) 2번 항목: 비가동 요인 (20~29행, I~R열 = 9~18번째 열 전체)
downtime_headers = [ws.cell(row=20, column=c).value for c in range(9, 19)]
downtime_rows = []
for r in range(21, 30):
    row_vals = [ws.cell(row=r, column=c).value for c in range(9, 19)]
    if any(v is not None and str(v).strip() != '' for v in row_vals):
        downtime_rows.append(row_vals)

# (4) 3번 항목: UPH / PPH 관리 (30~38행, B~Q열 = 2~17번째 열 전체)
uph_headers_r1 = [ws.cell(row=29, column=c).value for c in range(2, 18)]
uph_headers_r2 = [ws.cell(row=30, column=c).value for c in range(2, 18)]

uph_rows = []
for r in range(31, 39):
    row_vals = [ws.cell(row=r, column=c).value for c in range(2, 18)]
    if any(v is not None and str(v).strip() != '' for v in row_vals):
        uph_rows.append(row_vals)

# ----------------------------------------------------
# 3. 대시보드 시트 생성 및 정돈된 레이아웃 적용
# ----------------------------------------------------
if "Executive_Dashboard" in wb.sheetnames:
    del wb["Executive_Dashboard"]

dash = wb.create_sheet(title="Executive_Dashboard", index=0)
dash.views.sheetView[0].showGridLines = True

# 서식 스타일
font_title = Font(name="맑은 고딕", size=15, bold=True, color="1F497D")
font_section = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
font_header = Font(name="맑은 고딕", size=10, bold=True, color="333333")
font_data = Font(name="맑은 고딕", size=10)
font_kpi_num = Font(name="맑은 고딕", size=11, bold=True, color="1F497D")

fill_section = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
fill_highlight = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Title
dash.merge_cells("B2:Q2")
dash["B2"] = f"📊 [임원 보고용] {report_title} 요약 대시보드"
dash["B2"].font = font_title
dash["B2"].alignment = Alignment(vertical="center")

def create_section(ws, start_cell, end_cell, text):
    ws.merge_cells(f"{start_cell}:{end_cell}")
    cell = ws[start_cell]
    cell.value = text
    cell.font = font_section
    cell.fill = fill_section
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# --- 섹션 1: 핵심 생산 KPI ---
create_section(dash, "B4", "Q4", "1. 핵심 생산 KPI & 출고 요약")

headers_kpi = ["구분", "발주량", "생산계획", "CAPA계획", "생산실적", "달성률(%)", "투입시간 (계획/실적)", "완제품출고 (계획/실적)"]
for idx, h in enumerate(headers_kpi, start=2):
    c = dash.cell(row=5, column=idx, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border

kpi_values = [
    "월간 합계", order_qty, plan_qty, capa_qty, actual_qty,
    f"{achieve_rate:.1f}%",
    f"{plan_time:,.1f}h / {actual_time:,.1f}h",
    f"{ship_plan:,}건 / {ship_actual:,}건"
]

for idx, val in enumerate(kpi_values, start=2):
    c = dash.cell(row=6, column=idx)
    c.border = thin_border
    c.font = font_kpi_num if idx in [6, 7] else font_data
    if isinstance(val, (int, float)):
        c.value = val
        c.number_format = '#,##0'
        c.alignment = Alignment(horizontal="right", vertical="center")
    else:
        c.value = val
        c.alignment = Alignment(horizontal="center", vertical="center")
    if idx == 6:
        c.fill = fill_highlight

# --- 섹션 2: 비가동 요인 종합 분석 ---
create_section(dash, "B8", "Q8", "2. 비가동 요인 종합 분석")

for col_idx, h in enumerate(downtime_headers, start=2):
    c = dash.cell(row=9, column=col_idx, value=h or "")
    c.font = font_header
    c.fill = fill_header
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border

curr_row = 10
for r_data in downtime_rows:
    for col_idx, val in enumerate(r_data, start=2):
        c = dash.cell(row=curr_row, column=col_idx)
        c.border = thin_border
        c.font = font_data
        if isinstance(val, (int, float)):
            c.value = val
            c.number_format = '#,##0.0' if isinstance(val, float) else '#,##0'
            c.alignment = Alignment(horizontal="right", vertical="center")
        else:
            c.value = val if val is not None else ""
            c.alignment = Alignment(horizontal="center", vertical="center")
    curr_row += 1

# --- 섹션 3: 라인별 UPH / PPH 상세 관리 현황 (전체 컬럼) ---
uph_start = curr_row + 1
create_section(dash, f"B{uph_start}", f"Q{uph_start}", "3. 라인별 UPH / PPH 상세 관리 현황")

# 상단 2줄 헤더 매핑
header_r1 = uph_start + 1
header_r2 = uph_start + 2

for col_idx in range(2, 18):
    h1 = uph_headers_r1[col_idx-2]
    h2 = uph_headers_r2[col_idx-2]
    
    c1 = dash.cell(row=header_r1, column=col_idx, value=h1 or "")
    c2 = dash.cell(row=header_r2, column=col_idx, value=h2 or "")
    
    for c in (c1, c2):
        c.font = font_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

curr_row = header_r2 + 1
for r_data in uph_rows:
    for col_idx, val in enumerate(r_data, start=2):
        c = dash.cell(row=curr_row, column=col_idx)
        c.border = thin_border
        c.font = font_data
        
        if isinstance(val, float):
            c.value = round(val, 2)
            c.number_format = '#,##0.00'
            c.alignment = Alignment(horizontal="right", vertical="center")
        elif isinstance(val, int):
            c.value = val
            c.number_format = '#,##0'
            c.alignment = Alignment(horizontal="right", vertical="center")
        else:
            c.value = val or ""
            c.alignment = Alignment(horizontal="left" if col_idx == 2 else "center", vertical="center")
    curr_row += 1

# 열 너비 자동 맞춤
for col in dash.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    dash.column_dimensions[col_letter].width = max(max_len + 3, 11)

output_path = "▣_26년_월간_실적보고_대시보드_완성본.xlsx"
wb.save(output_path)
print(f"🎉 완벽히 정돈된 대시보드가 생성되었습니다: {output_path}")
